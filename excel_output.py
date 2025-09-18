import re
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill

class ExcelOutput:

    DIGITS_FIXED = 8 # how many digits to include in the floats
    BLINK_THRESHOLD = 2 # if pupil size is below this number it's considered a blink
    PUPIL_TS_GAP_MS = 14.5 # the approximate time in ms between pupil measurements
    BEFORE_EVENT_RANGE_MS = 90 # the range of values, in ms, that we want to account for when we measure pupil size before an event
    # vector to vf point conversion
    VEC_TO_VF = {
            '(-0.78, 0.78, 2.00)': 12,
            '(-0.15, 0.15, 2.00)': 33,
            '(0.78, 0.78, 2.00)': 17,
            '(0.15, 0.15, 2.00)': 34,
            '(0.78, -0.78, 2.00)': 65,
            '(0.15, -0.15, 2.00)': 44,
            '(-0.78, -0.78, 2.00)': 60,
            '(-0.15, -0.15, 2.00)': 43,
        }

    def write_to_excel(self, marker_streams, pupil_streams, results_dir, add_to_existing, output_filename, streams_filename):
        existing_path = os.path.join(results_dir, output_filename)

        if add_to_existing and not os.path.exists(existing_path):
            raise FileNotFoundError(f"Missing output path for dir {results_dir} and file {output_filename}")

        print(f"Writing to {existing_path}")
        subject_id, eeg_part = self.parse_filename(streams_filename)
        data = self.data_from_stream(marker_streams, pupil_streams, subject_id, eeg_part)
        headers = ["Subject", "EEG", "Label", "VF", "Avg Before", "Min Avg After", "%"]
        spans = [1, 2, 5, 1, 2, 2, 2]
        start_row = 2
        wb = None
        ws = None

        if add_to_existing:
            wb = load_workbook(existing_path)
            ws = wb.active
            start_row = ws.max_row + 1 # first empty row
        else:
            wb = Workbook()
            ws = wb.active

            col = 1
            for i, header in enumerate(headers):
                span = spans[i]
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color="E9EBF0", end_color="E9EBF0", fill_type="solid")
                col += span

        # fill data
        for r, row in enumerate(data, start_row):
            col = 1
            for i, value in enumerate(row):
                span = spans[i]
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + span - 1)
                # ws.cell(row=r, column=col, value=value)
                cell = ws.cell(row=r, column=col, value=value)
                cell.alignment = Alignment(horizontal="left")
                col += span

                if i == 6:
                    try:
                        if float(value) < -15:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    except (ValueError, TypeError):
                        pass

        wb.save(existing_path)
        print(f"Excel file saved to: {existing_path}")


    def data_from_stream(self, marker_streams, pupil_streams, subject_id, eeg_part):
        event_dict = self.build_event_dict(marker_streams, pupil_streams)

        data = []
        for event_name, info in event_dict.items():
            data.append([
                subject_id,
                eeg_part,
                info.get('label'),
                info.get('vf'),
                info.get('average_size_before'),
                info.get('average_size_after'),
                info.get('percentage_difference'),
            ])
        return data


    def build_event_dict(self, event_stream, pupil_stream):
        events = event_stream[0]['time_series']
        timestamps = event_stream[0]['time_stamps']

        pupil_dict = self.build_pupil_dict(pupil_stream)
        pupil_dict_keys = list(pupil_dict.keys())
        event_dict = {}

        for i in range(len(events)):
            event = events[i]
            ts = timestamps[i]
            event_name = self.clean_event_name(event[0])
            event_vector = self.get_vector_from_name(event_name)
            event_luminescence = self.get_luminescence_from_name(event_name)
            event_vf = self.vector_to_vf(event_vector)
            last_event_ts = timestamps[-1]
            fixed_digits_ts = round(ts, self.DIGITS_FIXED)
            fixed_digits_next_event_ts = round(timestamps[i + 1] if i + 1 < len(timestamps) else last_event_ts, self.DIGITS_FIXED)
            average_size_before = self.get_average_size_before_event(fixed_digits_ts, pupil_dict, pupil_dict_keys)

            if event_name not in event_dict:
                event_dict[event_name] = {
                    "start": fixed_digits_ts,
                    "stop": last_event_ts,
                    "vector": event_vector,
                    "luminescence": event_luminescence,
                    "vf": event_vf,
                    "label": event_name,
                    "average_size_before": average_size_before
                }
            else:
                event_dict[event_name]["stop"] = fixed_digits_ts
                event_dict[event_name]["next_event_start"] = fixed_digits_next_event_ts
                event_dict[event_name]["duration"] = self.calc_event_duration(event_dict[event_name].get("start"), fixed_digits_ts)

        for event in event_dict:
            event_end = round(event_dict[event].get("stop", event_stream[0]["time_stamps"][-1]), self.DIGITS_FIXED)
            next_start = round(event_dict[event].get("next_event_start", event_stream[0]["time_stamps"][-1]), self.DIGITS_FIXED)

            size_after = self.get_average_size_after_event(event_end, next_start, pupil_dict, pupil_dict_keys)
            event_dict[event]["average_size_after"] = size_after

            percentage_difference = self.calculate_percent_change(event_dict[event].get("average_size_before"), size_after)
            event_dict[event]["percentage_difference"] = percentage_difference

        return event_dict

    def build_pupil_dict(self, pupil_streams):

        pupil_sizes = np.array(pupil_streams[0]['time_series'])
        pupil_ts = np.array(pupil_streams[0]['time_stamps'])

        chosen_pupil = self.select_valid_pupil(pupil_sizes)
        pupil_dict = {}

        mask = [False] * len(chosen_pupil)
        for i, val in enumerate(chosen_pupil):
            if val == -1:
                for j in range(i - 2, i + 3):
                    if 0 <= j < len(mask):
                        mask[j] = True

        for time_stamp, pupil_size, is_blink in zip(pupil_ts, chosen_pupil, mask):
            fixed_time_stamp = round(time_stamp, self.DIGITS_FIXED)
            pupil_dict[fixed_time_stamp] = -1 if is_blink else pupil_size

        return pupil_dict

    def get_closest_timestamp(self, target_ts, pupil_dict_keys):
        idx = min(range(len(pupil_dict_keys)), key=lambda i: abs(pupil_dict_keys[i] - target_ts))
        return round(pupil_dict_keys[idx], self.DIGITS_FIXED), idx

    # get all timestamps between event_ts and end_ms_count from the pupil dict
    def get_timestamps_before_event(self, event_ts, pupil_dict, pupil_dict_keys):
        samples = []
        gap_s = self.PUPIL_TS_GAP_MS / 1000
        end_s_count = self.BEFORE_EVENT_RANGE_MS / 1000
        sampling_pointer, _ = self.get_closest_timestamp(event_ts, pupil_dict_keys)
        sampling_end, _ = self.get_closest_timestamp(event_ts + end_s_count, pupil_dict_keys)

        while sampling_pointer < sampling_end:
            ts, _ = self.get_closest_timestamp(sampling_pointer, pupil_dict_keys)
            pupil_size = pupil_dict[ts]
            samples.append(pupil_size)
            sampling_pointer = ts + gap_s

        return samples

    def get_average_size_before_event(self, event_start, pupil_dict, pupil_dict_keys):
        samples_before_start = self.get_timestamps_before_event(event_start, pupil_dict, pupil_dict_keys)
        average_pupil_size_before = sum(samples_before_start) / len(samples_before_start)
        return average_pupil_size_before

    def get_average_size_after_event(self, event_start, event_end, pupil_dict, pupil_dict_keys, count=3):
        # TODO: refactor to avoid max values
        samples = [1000] * count
        start_ts, start_i = self.get_closest_timestamp(event_start, pupil_dict_keys)
        end_ts, end_i = self.get_closest_timestamp(event_end, pupil_dict_keys)

        if event_start == event_end:
            ts, _ = self.get_closest_timestamp(event_start, pupil_dict_keys)
            return pupil_dict.get(ts)

        searching_range = pupil_dict_keys[start_i: end_i + 1]

        for ts in searching_range:
            pupil_size = pupil_dict.get(ts)
            if max(samples) > pupil_size > self.BLINK_THRESHOLD:
                idx = samples.index(max(samples))
                samples[idx] = pupil_size

        return sum(samples) / len(samples)

    def compile_filename(self):
        date_time = self.get_date_time()
        return f"Analytics-{date_time}.xlsx"

    @staticmethod
    def select_valid_pupil(pupil_data):
        if pupil_data.shape[1] >= 2:
            left = pupil_data[:, 0]
            right = pupil_data[:, 1]
            if not np.all(left == -1):
                return left
            elif not np.all(right == -1):
                return right
            else:
                raise ValueError("Both pupils are invalid.")
        else:
            pupil = pupil_data[:, 0]
            if np.all(pupil == -1):
                raise ValueError("Pupil data is invalid.")
            return pupil

    @staticmethod
    def get_date_time():
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H-%M-%S")
        return f"{date_str}_{time_str}"

    @staticmethod
    def parse_filename(filename):
        subject_id = None
        eeg_part = None

        try:
            match = re.search(r'sub-P(\d+)', filename)
            if match:
                subject_id = int(match.group(1))

            match = re.search(r'(eeg[^.]*)\.xdf$', filename)
            if match:
                eeg_part = match.group(1)
        except Exception:
            subject_id = filename
            eeg_part = filename

        return subject_id, eeg_part


    @staticmethod
    def clean_event_name(event_name):
        s = event_name.lower()
        s = re.sub(r'\b(start|stop)\b', '', s)
        s = re.sub(r'for\s+\d+(\.\d+)?\s+seconds', '', s)
        s = ' '.join(s.split())
        s = s.rstrip('.')
        return s

    @staticmethod
    def get_vector_from_name(event_name):
        match = re.search(r'\(([^)]+)\)', event_name)
        if match:
            return f"({match.group(1).strip()})"
        return None

    @staticmethod
    def get_luminescence_from_name(event_name):
        return event_name.split(" in (")[0].strip()

    @staticmethod
    def calc_event_duration(start_ts, stop_ts):
        duration = stop_ts - start_ts
        return f"{duration:.2f}"

    @staticmethod
    def calculate_percent_change(a, b):
        if a is None or b is None:
            return None
        return round(float((b - a) / a * 100), 3)

    def vector_to_vf(self, vec):
        vf = self.VEC_TO_VF.get(vec)
        if vf is None:
            return vec
        return vf


